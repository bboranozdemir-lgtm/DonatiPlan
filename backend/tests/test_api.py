import sys
import base64
from io import BytesIO
import tempfile
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from rebarflow.api import app
from rebarflow.store import SqliteStore


class ApiTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tempdir = tempfile.TemporaryDirectory()
        self.original_store = app.state.store
        app.state.store = SqliteStore(Path(self.tempdir.name) / "api-public.db")
        self.client = TestClient(app)

    def tearDown(self) -> None:
        self.client.close()
        app.state.store = self.original_store
        self.tempdir.cleanup()

    def test_health(self) -> None:
        response = self.client.get("/api/v1/health")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["status"], "ok")
        self.assertEqual(response.headers["x-content-type-options"], "nosniff")

    def test_cors_preflight_allows_local_frontend(self) -> None:
        response = self.client.options(
            "/api/v1/projects/example/inventory",
            headers={
                "Origin": "http://127.0.0.1:5173",
                "Access-Control-Request-Method": "PUT",
                "Access-Control-Request-Headers": "content-type",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.headers["access-control-allow-origin"],
            "http://127.0.0.1:5173",
        )

    def test_optimization_response_contains_cut_plan(self) -> None:
        response = self.client.post(
            "/api/v1/optimize",
            json={
                "demands": [
                    {
                        "mark": "K1",
                        "diameter_mm": 16,
                        "length_mm": 4200,
                        "quantity": 2,
                    },
                    {
                        "mark": "K2",
                        "diameter_mm": 16,
                        "length_mm": 3600,
                        "quantity": 1,
                    },
                ],
                "mode": "auto",
            },
        )
        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertEqual(body["solver_used"], "exact")
        self.assertEqual(body["summary"]["purchased_bar_count"], 1)
        self.assertEqual(body["summary"]["total_waste_mm"], 0)
        self.assertEqual(body["summary"]["real_scrap_mm"], 0)
        self.assertEqual(body["summary"]["waste_rate"], 0)
        self.assertEqual(body["summary"]["real_scrap_rate"], 0)
        self.assertEqual(body["reusable_remnants"], [])
        self.assertEqual(body["patterns"][0]["remaining_mm"], 0)

    def test_rejects_unknown_fields(self) -> None:
        response = self.client.post(
            "/api/v1/optimize",
            json={
                "demands": [
                    {
                        "mark": "K1",
                        "diameter_mm": 16,
                        "length_mm": 4200,
                        "quantity": 1,
                        "unexpected": True,
                    }
                ]
            },
        )
        self.assertEqual(response.status_code, 422)

    def test_rejects_length_over_stock(self) -> None:
        response = self.client.post(
            "/api/v1/optimize",
            json={
                "demands": [
                    {
                        "mark": "K1",
                        "diameter_mm": 16,
                        "length_mm": 13000,
                        "quantity": 1,
                    }
                ]
            },
        )
        self.assertEqual(response.status_code, 422)
        self.assertIn("exceeds", response.json()["detail"])

    def test_csv_import_reports_valid_and_invalid_rows(self) -> None:
        response = self.client.post(
            "/api/v1/import/csv",
            json={
                "content": "Poz;Çap;Boy;Adet\nK1;16;4200;2\nK2;xx;3000;1\n"
            },
        )
        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertEqual(body["accepted_count"], 1)
        self.assertEqual(body["rejected_count"], 1)

    def test_xlsx_template_can_be_downloaded_and_imported(self) -> None:
        template = self.client.get("/api/v1/templates/bbs.xlsx")
        self.assertEqual(template.status_code, 200)
        imported = self.client.post(
            "/api/v1/import/xlsx",
            json={
                "filename": "BBS.xlsx",
                "content_base64": base64.b64encode(template.content).decode("ascii"),
            },
        )
        self.assertEqual(imported.status_code, 200)
        self.assertEqual(imported.json()["accepted_count"], 4)


class PersistentApiTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tempdir = tempfile.TemporaryDirectory()
        self.original_store = app.state.store
        app.state.store = SqliteStore(Path(self.tempdir.name) / "api.db")
        self.client = TestClient(app)

    def tearDown(self) -> None:
        self.client.close()
        app.state.store = self.original_store
        self.tempdir.cleanup()

    def create_project(self) -> str:
        response = self.client.post(
            "/api/v1/projects",
            json={"name": "Gerçek Proje", "site": "İstanbul"},
        )
        self.assertEqual(response.status_code, 201)
        return response.json()["id"]

    def test_project_inventory_and_run_history_flow(self) -> None:
        project_id = self.create_project()

        demand_draft = self.client.put(
            f"/api/v1/projects/{project_id}/demands",
            json={
                "demands": [
                    {
                        "mark": "K1-TASLAK",
                        "diameter_mm": 16,
                        "length_mm": 3000,
                        "quantity": 1,
                        "phase": 1,
                    }
                ]
            },
        )
        self.assertEqual(demand_draft.status_code, 200)
        self.assertEqual(
            self.client.get(f"/api/v1/projects/{project_id}/demands").json()[0]["mark"],
            "K1-TASLAK",
        )

        inventory = self.client.put(
            f"/api/v1/projects/{project_id}/inventory",
            json={
                "items": [
                    {
                        "stock_code": "ART-16-01",
                        "diameter_mm": 16,
                        "length_mm": 3500,
                        "steel_grade": "B420C",
                    }
                ]
            },
        )
        self.assertEqual(inventory.status_code, 200)
        self.assertEqual(inventory.json()[0]["stock_code"], "ART-16-01")
        item_id = inventory.json()[0]["id"]
        qr = self.client.get(
            f"/api/v1/projects/{project_id}/inventory/{item_id}/qr.png"
        )
        self.assertEqual(qr.status_code, 200)
        self.assertTrue(qr.content.startswith(b"\x89PNG"))

        labels = self.client.get(
            f"/api/v1/projects/{project_id}/inventory-labels.pdf"
        )
        self.assertEqual(labels.status_code, 200)
        self.assertTrue(labels.content.startswith(b"%PDF"))

        optimized = self.client.post(
            f"/api/v1/projects/{project_id}/optimize",
            json={
                "demands": [
                    {
                        "mark": "=1+1",
                        "diameter_mm": 16,
                        "length_mm": 3000,
                        "quantity": 1,
                    }
                ],
                "remnants": [
                    {"id": "ART-16-01", "diameter_mm": 16, "length_mm": 3500}
                ],
                "mode": "auto",
            },
        )
        self.assertEqual(optimized.status_code, 200)
        self.assertIsNotNone(optimized.json()["run_id"])
        self.assertEqual(optimized.json()["summary"]["purchased_bar_count"], 0)
        self.assertGreater(optimized.json()["summary"]["demand_weight_kg"], 0)
        self.assertIsNotNone(optimized.json()["comparison"])

        runs = self.client.get(f"/api/v1/projects/{project_id}/runs")
        self.assertEqual(runs.status_code, 200)
        self.assertEqual(len(runs.json()), 1)
        self.assertEqual(runs.json()[0]["solver_used"], "exact")

        xlsx_report = self.client.get(
            f"/api/v1/projects/{project_id}/runs/{optimized.json()['run_id']}/report.xlsx"
        )
        self.assertEqual(xlsx_report.status_code, 200)
        self.assertTrue(xlsx_report.content.startswith(b"PK"))
        workbook = load_workbook(BytesIO(xlsx_report.content), data_only=False)
        self.assertEqual(workbook.sheetnames, ["Özet", "Talepler", "Kesim Planı", "Kullanılabilir Artık", "Mevcut Stok"])
        self.assertTrue(str(workbook["Özet"]["B20"].value).startswith("="))
        self.assertEqual(workbook["Özet"]["B21"].value, "=B20*'Talepler'!K2")
        self.assertEqual(workbook["Özet"]["B22"].value, "=B20*'Talepler'!K3")
        self.assertEqual(workbook["Talepler"]["A3"].value, "'=1+1")
        self.assertEqual(workbook["Kullanılabilir Artık"]["A2"].value, "Kaynak Stok")

        pdf_report = self.client.get(
            f"/api/v1/projects/{project_id}/runs/{optimized.json()['run_id']}/report.pdf"
        )
        self.assertEqual(pdf_report.status_code, 200)
        self.assertTrue(pdf_report.content.startswith(b"%PDF"))

        committed = self.client.post(
            f"/api/v1/projects/{project_id}/runs/{optimized.json()['run_id']}/commit"
        )
        self.assertEqual(committed.status_code, 200)
        self.assertEqual(committed.json()["consumed_remnant_count"], 1)

        inventory_after = self.client.get(
            f"/api/v1/projects/{project_id}/inventory"
        ).json()
        self.assertEqual(len(inventory_after), 0)
        movements = self.client.get(
            f"/api/v1/projects/{project_id}/movements"
        )
        self.assertEqual(movements.status_code, 200)
        movement_types = {item["movement_type"] for item in movements.json()}
        self.assertIn("inventory_added", movement_types)
        self.assertIn("consumed", movement_types)
        self.assertTrue(all(item["details"].get("actor") == "local" for item in movements.json()))
        self.assertIn("output_scrap", movement_types)

    def test_unknown_project_returns_404(self) -> None:
        response = self.client.get("/api/v1/projects/missing/inventory")
        self.assertEqual(response.status_code, 404)

    def test_project_settings_control_stock_and_kerf(self) -> None:
        project_id = self.create_project()
        settings = self.client.put(
            f"/api/v1/projects/{project_id}/settings",
            json={
                "stock_length_mm": 100,
                "min_reusable_mm": 10,
                "kerf_mm": 2,
                "steel_price_per_kg": 0,
                "carbon_kg_per_kg": 0,
                "currency": "TRY",
            },
        )
        self.assertEqual(settings.status_code, 200)

        optimized = self.client.post(
            f"/api/v1/projects/{project_id}/optimize",
            json={
                "demands": [
                    {
                        "mark": "A",
                        "diameter_mm": 16,
                        "length_mm": 49,
                        "quantity": 2,
                    }
                ]
            },
        )
        self.assertEqual(optimized.status_code, 200)
        self.assertEqual(optimized.json()["summary"]["purchased_bar_count"], 2)
        self.assertEqual(optimized.json()["summary"]["kerf_loss_mm"], 4)

    def test_project_backup_download_and_restore(self) -> None:
        project_id = self.create_project()
        backup = self.client.get(f"/api/v1/projects/{project_id}/backup.json")
        self.assertEqual(backup.status_code, 200)
        self.assertEqual(backup.json()["format"], "rebarflow-project-backup")

        restored = self.client.post(
            "/api/v1/backups/restore",
            json={"backup": backup.json()},
        )
        self.assertEqual(restored.status_code, 201)
        self.assertNotEqual(restored.json()["id"], project_id)

    def test_qr_inventory_transition_endpoint(self) -> None:
        project_id = self.create_project()
        inventory = self.client.put(
            f"/api/v1/projects/{project_id}/inventory",
            json={
                "items": [
                    {"stock_code": "QR-API", "diameter_mm": 12, "length_mm": 2200}
                ]
            },
        ).json()
        item_id = inventory[0]["id"]
        transitioned = self.client.post(
            f"/api/v1/projects/{project_id}/inventory/{item_id}/transition",
            json={"target_status": "reserved", "note": "Döşeme D1"},
        )
        self.assertEqual(transitioned.status_code, 200)
        self.assertEqual(transitioned.json()["status"], "reserved")


class AuthenticationApiTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tempdir = tempfile.TemporaryDirectory()
        self.original_store = app.state.store
        app.state.store = SqliteStore(Path(self.tempdir.name) / "auth.db")
        self.client = TestClient(app)

    def tearDown(self) -> None:
        self.client.close()
        app.state.store = self.original_store
        self.tempdir.cleanup()

    def test_auth_endpoints_remain_but_api_is_local_access(self) -> None:
        status = self.client.get("/api/v1/auth/status").json()
        self.assertTrue(status["setup_required"])

        bootstrap = self.client.post(
            "/api/v1/auth/bootstrap",
            json={"username": "admin", "password": "cok-guclu-parola"},
        )
        self.assertEqual(bootstrap.status_code, 201)
        self.assertEqual(bootstrap.json()["role"], "admin")
        self.assertIn("HttpOnly", bootstrap.headers["set-cookie"])
        self.assertIn("SameSite=lax", bootstrap.headers["set-cookie"])

        viewer = self.client.post(
            "/api/v1/users",
            json={
                "username": "izleyici",
                "password": "izleyici-parola-123",
                "role": "viewer",
            },
        )
        self.assertEqual(viewer.status_code, 201)

        anonymous = TestClient(app)
        self.assertEqual(anonymous.get("/api/v1/projects").status_code, 200)
        login_response = anonymous.post(
            "/api/v1/auth/login",
            json={"username": "izleyici", "password": "izleyici-parola-123"},
        )
        self.assertEqual(login_response.status_code, 200)
        self.assertEqual(anonymous.get("/api/v1/projects").status_code, 200)
        self.assertEqual(anonymous.get("/api/v1/users").status_code, 200)
        created = anonymous.post(
            "/api/v1/projects",
            json={"name": "Yerel Proje"},
        )
        self.assertEqual(created.status_code, 201)
        self.assertEqual(anonymous.post("/api/v1/auth/logout").status_code, 200)
        self.assertEqual(anonymous.get("/api/v1/projects").status_code, 200)
        anonymous.close()


if __name__ == "__main__":
    unittest.main()
