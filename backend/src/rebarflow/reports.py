from __future__ import annotations

import json
from html import escape
from io import BytesIO
from pathlib import Path

import qrcode
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.pdfgen import canvas

from .store import OptimizationRunRecord, ProjectRecord, ProjectSettingsRecord, RemnantRecord


class ReportBuilder:
    DARK = "153D31"
    LIME = "D8FF69"
    PALE = "EAF0EC"
    GRID = "DDE3DE"

    @staticmethod
    def _excel_text(value: object) -> str:
        text = str(value)
        return f"'{text}" if text.startswith(("=", "+", "-", "@")) else text

    def build_xlsx(
        self,
        project: ProjectRecord,
        settings: ProjectSettingsRecord,
        run: OptimizationRunRecord,
        inventory: list[RemnantRecord],
    ) -> bytes:
        workbook = Workbook()
        summary = workbook.active
        summary.title = "Özet"
        demands = workbook.create_sheet("Talepler")
        cutting = workbook.create_sheet("Kesim Planı")
        reusable = workbook.create_sheet("Kullanılabilir Artık")
        stock = workbook.create_sheet("Mevcut Stok")
        summary_data = run.result_data["summary"]
        total_waste_mm = int(
            summary_data.get(
                "total_waste_mm",
                summary_data.get("reusable_output_mm", 0) + summary_data.get("scrap_output_mm", 0),
            )
        )
        used_source_length_mm = int(
            summary_data.get(
                "used_source_length_mm",
                summary_data.get("purchased_length_mm", 0) + summary_data.get("remnant_source_used_mm", 0),
            )
        )
        new_stock_waste_mm = int(summary_data.get("new_stock_waste_mm", total_waste_mm))
        real_scrap_mm = int(summary_data.get("real_scrap_mm", summary_data.get("scrap_output_mm", 0)))
        waste_rate = float(summary_data.get("waste_rate", run.purchase_waste_rate))
        new_stock_waste_rate = float(summary_data.get("new_stock_waste_rate", 0))
        new_stock_waste_rate_display: float | str = (
            new_stock_waste_rate
            if int(summary_data.get("purchased_length_mm", 0)) > 0
            else "Yeni stok kullanılmadı"
        )
        real_scrap_rate = float(summary_data.get("real_scrap_rate", 0))

        self._title(summary, "A1:F1", "DonatıPlan · Optimizasyon Raporu · by BBO")
        summary.append([])
        summary.append([
            "Proje",
            self._excel_text(project.name),
            "Şantiye",
            self._excel_text(project.site),
            "Çalışma",
            run.id[:8],
        ])
        summary.append(["Tarih", run.created_at, "Çözücü", run.solver_used, "Durum", run.status])
        summary.append([])
        summary.append(["Gösterge", "Değer", "Birim", "Açıklama"])
        summary.append(["Talep parçası", run.piece_count, "adet", "Toplam kesilecek parça"])
        summary.append(["Toplam ihtiyaç uzunluğu", summary_data["demand_length_mm"] / 1000, "m", "Kesilecek net donatı boyu"])
        summary.append(["Kullanılan stok çubuk sayısı", run.purchased_bar_count, "adet", "Satın alınacak standart çubuk"])
        summary.append(["Satın alınan toplam stok", summary_data["purchased_length_mm"] / 1000, "m", "Yeni stok toplamı"])
        summary.append(["Kullanılan toplam kaynak", used_source_length_mm / 1000, "m", "Yeni stok + kullanılan mevcut artık"])
        summary.append(["Yeni stok artık uzunluğu", new_stock_waste_mm / 1000, "m", "Sadece satın alınan yeni stoktan kalan boy"])
        summary.append(["Yeni stok artık oranı", new_stock_waste_rate_display, "%", "Yeni stok artığı / satın alınan yeni stok"])
        summary.append(["Toplam artık uzunluk", total_waste_mm / 1000, "m", "Kesimlerden sonra kalan toplam boy"])
        summary.append(["İşlem toplam artık oranı", waste_rate, "%", "Tüm kalan artık / kullanılan toplam kaynak"])
        summary.append(["Gerçek hurda uzunluğu", real_scrap_mm / 1000, "m", f"{settings.min_reusable_mm} mm altı kalan parçalar"])
        summary.append(["Gerçek hurda oranı", real_scrap_rate, "%", "Gerçek hurda / kullanılan toplam kaynak"])
        summary.append(["Kullanılabilir artık uzunluğu", summary_data["reusable_output_mm"] / 1000, "m", "Stokta tekrar kullanılabilir kalan boy"])
        summary.append(["Kesim kaybı", summary_data.get("kerf_loss_mm", 0) / 1000, "m", "Testere/makas ağzı kaybı"])
        summary.append(["Tahmini donatı ağırlığı", "='Talepler'!H2", "kg", "Ø²/162 yaklaşımı"])
        summary.append(["Tahmini malzeme maliyeti", "=B20*'Talepler'!K2", settings.currency, "Ayarlar sayfasındaki kg fiyatı"])
        summary.append(["Tahmini karbon", "=B20*'Talepler'!K3", "kgCO2e", "Proje karbon katsayısı"])

        summary["B13"].number_format = "0.0%"
        summary["B15"].number_format = "0.0%"
        summary["B17"].number_format = "0.0%"
        summary["B20"].number_format = "#,##0.00"
        summary["B21"].number_format = "#,##0.00"
        summary["B22"].number_format = "#,##0.00"
        self._style_table(summary, "A6:D22")
        summary.freeze_panes = "A6"

        self._title(demands, "A1:H1", "Donatı Talepleri")
        demands.append(["Poz", "Çap (mm)", "Boy (mm)", "Adet", "Faz", "Toplam Boy (m)", "Birim kg/m", "Toplam kg"])
        request_demands = run.request_data.get("demands", [])
        start_row = 3
        for index, item in enumerate(request_demands, start=start_row):
            demands.append([
                self._excel_text(item["mark"]),
                item["diameter_mm"],
                item["length_mm"],
                item["quantity"],
                item.get("phase", 0),
                f"=C{index}*D{index}/1000",
                f"=B{index}^2/162",
                f"=F{index}*G{index}",
            ])
        total_row = max(start_row, start_row + len(request_demands))
        demands.cell(total_row, 7, "Toplam")
        demands.cell(total_row, 8, f"=SUM(H{start_row}:H{total_row - 1})" if request_demands else 0)
        summary["B20"] = f"='Talepler'!H{total_row}"
        demands.cell(2, 10, "Çelik fiyatı/kg")
        demands.cell(2, 11, settings.steel_price_per_kg)
        demands.cell(3, 10, "Karbon kgCO2e/kg")
        demands.cell(3, 11, settings.carbon_kg_per_kg)
        self._style_table(demands, f"A2:H{total_row}")
        demands.freeze_panes = "A3"

        self._title(cutting, "A1:H1", "Çubuk Bazlı Kesim Planı")
        cutting.append(["Stok Kodu", "Kaynak", "Çap", "Stok Boyu", "Kesimler", "Kesim Kaybı", "Kalan", "Durum"])
        for pattern in run.result_data.get("patterns", []):
            remaining = int(pattern.get("remaining_mm", 0))
            cutting.append([
                self._excel_text(pattern["stock_id"]),
                "Yeni" if pattern["source"] == "new" else "Artık",
                pattern["diameter_mm"],
                pattern["stock_length_mm"],
                self._excel_text(
                    " + ".join(f"{cut['mark']} ({cut['length_mm']})" for cut in pattern["cuts"])
                ),
                pattern.get("kerf_loss_mm", 0),
                remaining,
                "Kullanılabilir" if remaining >= settings.min_reusable_mm else ("Tam kullanıldı" if remaining == 0 else "Hurda"),
            ])
        self._style_table(cutting, f"A2:H{max(2, cutting.max_row)}")
        cutting.freeze_panes = "A3"

        self._title(reusable, "A1:E1", "Kullanılabilir Artık Listesi")
        reusable.append(["Kaynak Stok", "Çap", "Boy", "Kaynak", "Açıklama"])
        reusable_rows = run.result_data.get("reusable_remnants", [])
        if not reusable_rows:
            reusable_rows = [
                {
                    "source_stock_id": pattern["stock_id"],
                    "diameter_mm": pattern["diameter_mm"],
                    "length_mm": pattern.get("remaining_mm", 0),
                    "source": pattern["source"],
                    "note": "Bu parça bu projede kullanılmadı; stokta tekrar kullanılabilir.",
                }
                for pattern in run.result_data.get("patterns", [])
                if int(pattern.get("remaining_mm", 0)) >= settings.min_reusable_mm
                and int(pattern.get("remaining_mm", 0)) > 0
            ]
        for item in reusable_rows:
            reusable.append([
                self._excel_text(item["source_stock_id"]),
                item["diameter_mm"],
                item["length_mm"],
                "Yeni" if item["source"] == "new" else "Artık",
                self._excel_text(item["note"]),
            ])
        self._style_table(reusable, f"A2:E{max(2, reusable.max_row)}")
        reusable.freeze_panes = "A3"

        self._title(stock, "A1:F1", "Mevcut Kullanılabilir Artık Stoku")
        stock.append(["Stok Kodu", "Çap", "Boy", "Çelik Sınıfı", "Durum", "Güncelleme"])
        for item in inventory:
            stock.append([
                self._excel_text(item.stock_code),
                item.diameter_mm,
                item.length_mm,
                self._excel_text(item.steel_grade),
                self._excel_text(item.status),
                item.updated_at,
            ])
        self._style_table(stock, f"A2:F{max(2, stock.max_row)}")
        stock.freeze_panes = "A3"

        for sheet in workbook.worksheets:
            sheet.sheet_view.showGridLines = False
            self._fit_columns(sheet)
            sheet.auto_filter.ref = sheet.dimensions if sheet.max_row > 1 else None

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def build_pdf(
        self,
        project: ProjectRecord,
        settings: ProjectSettingsRecord,
        run: OptimizationRunRecord,
    ) -> bytes:
        output = BytesIO()
        font_name = self._register_font()
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "RFTitle",
            parent=styles["Title"],
            fontName=font_name,
            fontSize=20,
            leading=24,
            textColor=colors.HexColor("#153D31"),
            alignment=TA_LEFT,
        )
        body_style = ParagraphStyle(
            "RFBody",
            parent=styles["BodyText"],
            fontName=font_name,
            fontSize=8,
            leading=11,
        )
        document = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            leftMargin=14 * mm,
            rightMargin=14 * mm,
            topMargin=12 * mm,
            bottomMargin=12 * mm,
            title=f"DonatıPlan - {project.name}",
        )
        summary = run.result_data["summary"]
        total_waste_mm = int(
            summary.get(
                "total_waste_mm",
                summary.get("reusable_output_mm", 0) + summary.get("scrap_output_mm", 0),
            )
        )
        used_source_length_mm = int(
            summary.get(
                "used_source_length_mm",
                summary.get("purchased_length_mm", 0) + summary.get("remnant_source_used_mm", 0),
            )
        )
        new_stock_waste_mm = int(summary.get("new_stock_waste_mm", total_waste_mm))
        real_scrap_mm = int(summary.get("real_scrap_mm", summary.get("scrap_output_mm", 0)))
        waste_rate = float(summary.get("waste_rate", run.purchase_waste_rate))
        new_stock_waste_rate = float(summary.get("new_stock_waste_rate", 0))
        new_stock_waste_rate_display = (
            f"%{new_stock_waste_rate * 100:.1f}"
            if int(summary.get("purchased_length_mm", 0)) > 0
            else "Yeni stok kullanılmadı"
        )
        real_scrap_rate = float(summary.get("real_scrap_rate", 0))
        story = [
            Paragraph("DonatıPlan · Akıllı Donatı Kesim, Artık ve Hurda Analizi · by BBO", title_style),
            Paragraph(
                f"Proje: {escape(project.name)} &nbsp;&nbsp; Şantiye: {escape(project.site or '-')} &nbsp;&nbsp; Çalışma: {run.id[:8]} &nbsp;&nbsp; Tarih: {run.created_at}",
                body_style,
            ),
            Spacer(1, 7 * mm),
        ]
        metric_data = [
            [
                "İhtiyaç",
                "Yeni Çubuk",
                "Yeni Stok",
                "Kull. Kaynak",
                "Yeni Stok Artık",
                "Yeni Stok Artık %",
                "Toplam Artık",
                "İşlem Artık %",
                "Gerçek Hurda",
                "Hurda Oranı",
                "Kull. Artık",
            ],
            [
                f"{summary['demand_length_mm'] / 1000:.2f} m",
                str(run.purchased_bar_count),
                f"{summary['purchased_length_mm'] / 1000:.2f} m",
                f"{used_source_length_mm / 1000:.2f} m",
                f"{new_stock_waste_mm / 1000:.2f} m",
                new_stock_waste_rate_display,
                f"{total_waste_mm / 1000:.2f} m",
                f"%{waste_rate * 100:.1f}",
                f"{real_scrap_mm / 1000:.2f} m",
                f"%{real_scrap_rate * 100:.1f}",
                f"{summary['reusable_output_mm'] / 1000:.2f} m",
            ],
        ]
        metric_table = Table(metric_data, colWidths=[23 * mm] * 11)
        metric_table.setStyle(self._pdf_table_style(font_name, header=True))
        story.extend([metric_table, Spacer(1, 8 * mm)])

        reusable_rows = run.result_data.get("reusable_remnants", [])
        if not reusable_rows:
            reusable_rows = [
                {
                    "source_stock_id": pattern["stock_id"],
                    "diameter_mm": pattern["diameter_mm"],
                    "length_mm": pattern.get("remaining_mm", 0),
                    "source": pattern["source"],
                    "note": "Bu parça bu projede kullanılmadı; stokta tekrar kullanılabilir.",
                }
                for pattern in run.result_data.get("patterns", [])
                if int(pattern.get("remaining_mm", 0)) >= settings.min_reusable_mm
                and int(pattern.get("remaining_mm", 0)) > 0
            ]
        if reusable_rows:
            reusable_table = Table(
                [["Kullanılabilir Artık", "Ø", "Boy", "Açıklama"]]
                + [
                    [
                        str(item["source_stock_id"]),
                        str(item["diameter_mm"]),
                        f"{int(item['length_mm'])} mm",
                        Paragraph(escape(str(item["note"])), body_style),
                    ]
                    for item in reusable_rows
                ],
                repeatRows=1,
                colWidths=[45 * mm, 12 * mm, 24 * mm, 160 * mm],
            )
            reusable_table.setStyle(self._pdf_table_style(font_name, header=True))
            story.extend([reusable_table, Spacer(1, 8 * mm)])

        pattern_rows = [["Stok", "Kaynak", "Ø", "Stok", "Kesimler (mm)", "Kayıp", "Kalan", "Durum"]]
        for pattern in run.result_data.get("patterns", []):
            remaining = int(pattern.get("remaining_mm", 0))
            cuts = " + ".join(
                f"{escape(str(cut['mark']))}:{cut['length_mm']}" for cut in pattern["cuts"]
            )
            status = "Kullanılabilir" if remaining >= settings.min_reusable_mm else ("Tam" if remaining == 0 else "Hurda")
            pattern_rows.append([
                str(pattern["stock_id"]),
                "Yeni" if pattern["source"] == "new" else "Artık",
                str(pattern["diameter_mm"]),
                str(pattern["stock_length_mm"]),
                Paragraph(cuts, body_style),
                str(pattern.get("kerf_loss_mm", 0)),
                str(remaining),
                status,
            ])
        plan_table = Table(
            pattern_rows,
            repeatRows=1,
            colWidths=[31 * mm, 17 * mm, 10 * mm, 18 * mm, 112 * mm, 16 * mm, 18 * mm, 25 * mm],
        )
        plan_table.setStyle(self._pdf_table_style(font_name, header=True))
        story.append(plan_table)
        document.build(story)
        return output.getvalue()

    def build_inventory_qr_png(self, project: ProjectRecord, item: RemnantRecord) -> bytes:
        payload = json.dumps(
            {
                "v": 1,
                "type": "rebarflow-remnant",
                "project_id": project.id,
                "item_id": item.id,
                "stock_code": item.stock_code,
            },
            ensure_ascii=False,
            separators=(",", ":"),
        )
        qr = qrcode.QRCode(version=None, box_size=8, border=2)
        qr.add_data(payload)
        qr.make(fit=True)
        image = qr.make_image(fill_color="#153D31", back_color="white")
        output = BytesIO()
        image.save(output, format="PNG")
        return output.getvalue()

    def build_inventory_labels_pdf(
        self,
        project: ProjectRecord,
        inventory: list[RemnantRecord],
    ) -> bytes:
        output = BytesIO()
        font_name = self._register_font()
        page_width, page_height = A4
        pdf = canvas.Canvas(output, pagesize=A4)
        pdf.setTitle(f"DonatıPlan QR Etiketleri - {project.name}")

        card_width = 90 * mm
        card_height = 55 * mm
        margin_x = 10 * mm
        margin_y = 10 * mm
        gap_x = 7 * mm
        gap_y = 5 * mm
        columns = 2
        rows = 5

        for index, item in enumerate(inventory):
            slot = index % (columns * rows)
            if index > 0 and slot == 0:
                pdf.showPage()
            column = slot % columns
            row = slot // columns
            x = margin_x + column * (card_width + gap_x)
            y = page_height - margin_y - (row + 1) * card_height - row * gap_y

            pdf.setStrokeColor(colors.HexColor("#AAB8B0"))
            pdf.setLineWidth(0.6)
            pdf.roundRect(x, y, card_width, card_height, 3 * mm, stroke=1, fill=0)
            pdf.setFillColor(colors.HexColor("#153D31"))
            pdf.rect(x, y + card_height - 10 * mm, card_width, 10 * mm, stroke=0, fill=1)
            pdf.setFillColor(colors.white)
            pdf.setFont(font_name, 9)
            pdf.drawString(x + 4 * mm, y + card_height - 6.5 * mm, "DonatıPlan · Artık Donatı")

            qr_bytes = self.build_inventory_qr_png(project, item)
            pdf.drawImage(
                ImageReader(BytesIO(qr_bytes)),
                x + 4 * mm,
                y + 4 * mm,
                width=35 * mm,
                height=35 * mm,
                preserveAspectRatio=True,
                mask="auto",
            )
            text_x = x + 43 * mm
            pdf.setFillColor(colors.HexColor("#153D31"))
            pdf.setFont(font_name, 11)
            pdf.drawString(text_x, y + 34 * mm, item.stock_code[:24])
            pdf.setFont(font_name, 9)
            pdf.drawString(text_x, y + 26 * mm, f"Çap: Ø{item.diameter_mm} mm")
            pdf.drawString(text_x, y + 20 * mm, f"Boy: {item.length_mm} mm")
            pdf.drawString(text_x, y + 14 * mm, f"Sınıf: {item.steel_grade}")
            pdf.setFillColor(colors.HexColor("#607169"))
            pdf.setFont(font_name, 7)
            pdf.drawString(text_x, y + 6 * mm, project.name[:28])

        if not inventory:
            pdf.setFont(font_name, 12)
            pdf.drawString(20 * mm, page_height - 25 * mm, "Etiketlenecek kullanılabilir artık parça bulunamadı.")
        pdf.save()
        return output.getvalue()

    def _title(self, sheet, merged_range: str, text: str) -> None:
        sheet.merge_cells(merged_range)
        cell = sheet[merged_range.split(":")[0]]
        cell.value = text
        cell.fill = PatternFill("solid", fgColor=self.DARK)
        cell.font = Font(color="FFFFFF", bold=True, size=16)
        cell.alignment = Alignment(vertical="center")
        sheet.row_dimensions[cell.row].height = 30

    def _style_table(self, sheet, range_ref: str) -> None:
        cells = sheet[range_ref]
        thin = Side(style="thin", color=self.GRID)
        for row_index, row in enumerate(cells):
            for cell in row:
                cell.border = Border(bottom=thin)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if row_index == 0:
                    cell.fill = PatternFill("solid", fgColor=self.LIME)
                    cell.font = Font(color=self.DARK, bold=True)

    @staticmethod
    def _fit_columns(sheet) -> None:
        for column_cells in sheet.columns:
            letter = get_column_letter(column_cells[0].column)
            maximum = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[letter].width = min(max(maximum + 2, 10), 48)

    @staticmethod
    def _register_font() -> str:
        candidates = [
            Path("C:/Windows/Fonts/arial.ttf"),
            Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
        ]
        for candidate in candidates:
            if candidate.exists():
                try:
                    pdfmetrics.registerFont(TTFont("RebarFlowFont", str(candidate)))
                    return "RebarFlowFont"
                except Exception:
                    continue
        return "Helvetica"

    @staticmethod
    def _pdf_table_style(font_name: str, header: bool) -> TableStyle:
        commands = [
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#DDE3DE")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7F4")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]
        if header:
            commands.extend([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D8FF69")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#153D31")),
                ("FONTNAME", (0, 0), (-1, 0), font_name),
            ])
        return TableStyle(commands)
